# -*- coding: utf-8 -*-  

import os,sys
os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.abspath('..'))
from jxnuJWspider import SearchClient
import openpyxl
import time,re

#==================================================
#
#  在本例中，我们将尝试教务在线中的 所在单位 查询
#  通过对物理学院20、19级各个班级的单位查询，
#  我们可以获取物理学院20、19级全体名单，并把结果保存在新建的表格中
#
#==================================================

if __name__ == "__main__":
    #创建查询会话 (需要你手动填入自己的教务在线的账号和密码)
    jwsc=SearchClient('202066601001','my password')

    #创建用于写入结果的表格
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "查询结果"
    ws.append(['学院','班级名称','学号','姓名','性别'])

    #获取物理学院班级列表(必须输入教务在线上的学院全称)
    college='物理与通信电子学院'
    classes=jwsc.getClasses(college)
    
    #你可以通过 jwsc.getUnits() 获取所有学院的全称列表
    """ 例：
    units = jwsc.getUnits()
    for u in units:
        print(u)
    """
    
    #使用re模块删选班级，去除以非20、19开头的班级
    print('删除前:',classes)
    for c in classes[::-1]:
        if not re.match('20|19',c):
            classes.remove(c)
    print('删除后:',classes)

    #=========================开始查询=========================
    i=1
    for c in classes:
        print(c)

        #调用查询并返回到data
        data=jwsc.searchByUnit('学生',college,c)

        #将数据写入表格
        for d in data:
            ws.cell(row=i+1,column=1).value = d['所在单位']
            ws.cell(row=i+1,column=2).value = d['班级名称']
            ws.cell(row=i+1,column=3).value = d['学号']
            ws.cell(row=i+1,column=4).value = d['姓名']
            ws.cell(row=i+1,column=5).value = d['性别']
            i+=1

        #设定一个延迟，防止批量访问对教务在线产生负担
        time.sleep(0.5)
    #=========================查询结束=========================

    #保存表格
    wb.save('例3查询结果.xlsx')
    print('查询完成')