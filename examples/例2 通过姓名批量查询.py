# -*- coding: utf-8 -*-  

import os
import sys
os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.abspath('..'))
from jxnuJWspider import SearchClient
import openpyxl
import time
import random

#==================================================
#
#  在本例中，我们通过爬虫进行批量查询，这将体现爬虫查询的真正优势
#  通过姓名精确查询，当出现同名同姓的情况时，查询会返回多个数据
#  我们将对同名同姓的情况进行简单处理(填充颜色)，最后把结果保存在新建的表格中
#
#==================================================

if __name__ == "__main__":
    #创建查询会话 (需要你手动填入自己的教务在线的账号和密码)
    jwsc=SearchClient('202066601001','my password')
    
    #打开要查询的源数据作为 表1
    wb1 = openpyxl.load_workbook('例2源数据.xlsx')
    ws1 = wb1[wb1.sheetnames[0]]

    #创建用于写入结果的 表2
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "查询结果"
    ws2.append(['学院','学号','姓名','性别','班级名称'])

    #选择几个颜色用于填充
    colors=['EA7272','E6B176','E0E379','92DB81','83D9D5','87AAD5','837FDD','B282DA','6EEE86']

    #=========================开始查询=========================
    i=1
    #从表1中依次取数据
    for r1 in ws1.values:
        print(i,r1[0])

        #调用查询并返回到data
        data=jwsc.search('学生',r1[0],'姓名','精确')

        #当查询值大于1时(同名同姓)，选取一种颜色作为填充，否则填充白色
        if len(data)>1:
            cfill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=colors[random.randint(0,8)])
        else:
            cfill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FFFFFF")

        #将数据写入表2，并填充颜色
        for d in data:
            ws2.cell(row=i+1,column=1).value = d['所在单位']
            ws2.cell(row=i+1,column=2).value = d['学号']
            ws2.cell(row=i+1,column=3).value = d['姓名']
            ws2.cell(row=i+1,column=4).value = d['性别']
            ws2.cell(row=i+1,column=5).value = d['班级名称']
            ws2.cell(row=i+1,column=3).fill=cfill
            i+=1

        #设定一个延迟，防止批量访问对教务在线产生负担
        time.sleep(0.5)
    #=========================查询结束=========================

    #保存表格
    wb2.save('例2查询结果.xlsx')
    print('查询完成')