# -*- coding: utf-8 -*-  

import os
import sys
os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.abspath('..'))
from jxnuJWspider import SearchClient
import openpyxl
import time

#===================================
#
#  在本例中，我们通过爬虫模糊查询一段教工号(00111)
#  这次查询会返回多个数据
#  我们把它保存在新建的表格中
#
#===================================

if __name__ == "__main__":
    #创建查询会话 (需要你手动填入自己的教务在线的账号和密码)
    jwsc=SearchClient('202066601001','my password')

    #查询并返回结果
    data=jwsc.search('教工','00111','教号','模糊')

    #创建表格，用于储存查询结果
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "查询结果"

    #写入表头
    ws.append(['教号','姓名','性别','所在单位'])

    #向表格中写入查询结果
    for i,d in enumerate(data):
        print(i,d['教号'])
        ws.cell(row=i+2,column=1).value = d['教号']
        ws.cell(row=i+2,column=2).value = d['姓名']
        ws.cell(row=i+2,column=3).value = d['性别']
        ws.cell(row=i+2,column=4).value = d['所在单位']
        
    #保存表格
    wb.save('例1查询结果.xlsx')
    print('查询完成')